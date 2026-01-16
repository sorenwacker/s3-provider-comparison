"""Typer CLI for S3 provider benchmarking."""

import csv
import json
import socket
from datetime import datetime
from pathlib import Path
from typing import Annotated, Optional

import typer
from openpyxl import Workbook, load_workbook
from rich.console import Console
from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn, TaskProgressColumn
from rich.table import Table

from s3bench import config as cfg
from s3bench.config import ProviderType
from s3bench.benchmark import run_benchmark, DEFAULT_ITERATIONS, BenchmarkResult
from s3bench.providers import create_provider
from s3bench.features import run_feature_tests, FEATURE_TESTS, FEATURE_DESCRIPTIONS, FeatureStatus
from s3bench.rclone import create_rclone_provider, check_rclone_installed
from s3bench.s5cmd import create_s5cmd_provider, check_s5cmd_installed


def _parse_provider_method(name: str) -> tuple[str, str]:
    """Parse provider name into (base_provider, method)."""
    if name.endswith("_rclone"):
        return name[:-7], "rclone"
    elif name.endswith("_s5cmd"):
        return name[:-6], "s5cmd"
    return name, "sdk"


def _format_mean_std(values: list[float], decimals: int = 2, std_decimals: int = None) -> str:
    """Format values as mean +/- std."""
    import statistics
    if not values:
        return "-"
    if std_decimals is None:
        std_decimals = decimals
    mean = statistics.mean(values)
    if len(values) > 1:
        std = statistics.stdev(values)
        return f"{mean:.{decimals}f}+/-{std:.{std_decimals}f}"
    return f"{mean:.{decimals}f}"


def get_results_dir() -> Path:
    """Get the results storage directory."""
    results_dir = Path.home() / ".config" / "s3bench" / "results"
    results_dir.mkdir(parents=True, exist_ok=True)
    return results_dir


def save_results(result: BenchmarkResult) -> Path:
    """Save benchmark results to JSON file."""
    results_dir = get_results_dir()
    filename = f"benchmark_{result.timestamp.strftime('%Y-%m-%d_%H-%M-%S')}.json"
    filepath = results_dir / filename

    data = {
        "timestamp": result.timestamp.isoformat(),
        "providers": {},
    }
    for provider_name, provider_result in result.provider_results.items():
        data["providers"][provider_name] = {
            "size_results": {
                str(size): {
                    "size_bytes": sr.size_bytes,
                    "upload_throughputs": sr.upload_throughputs,
                    "download_throughputs": sr.download_throughputs,
                    "latencies": sr.latencies,
                }
                for size, sr in provider_result.size_results.items()
            },
            "errors": provider_result.errors,
        }

    with open(filepath, "w") as f:
        json.dump(data, f, indent=2)

    return filepath


def get_machine_info() -> dict:
    """Get current machine information."""
    hostname = socket.gethostname()

    # Get local IP by connecting to a remote address (doesn't actually send data)
    ip_address = "unknown"
    try:
        with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as s:
            s.connect(("8.8.8.8", 80))
            ip_address = s.getsockname()[0]
    except Exception:
        try:
            ip_address = socket.gethostbyname(hostname)
        except socket.gaierror:
            pass

    return {"hostname": hostname, "ip_address": ip_address}


def get_csv_path() -> Path:
    """Get the CSV results file path."""
    return get_results_dir() / "benchmarks.csv"


def save_results_csv(result: BenchmarkResult) -> Path:
    """Append benchmark results to CSV file in tidy/long format with individual iterations."""
    csv_path = get_csv_path()
    machine_info = get_machine_info()
    file_exists = csv_path.exists()

    # Collect all rows in tidy format (one row per metric per iteration)
    rows = []
    for provider_name, provider_result in result.provider_results.items():
        # Parse provider name and method
        if provider_name.endswith("_rclone"):
            base_provider = provider_name[:-7]  # Remove "_rclone"
            method = "rclone"
        elif provider_name.endswith("_s5cmd"):
            base_provider = provider_name[:-6]  # Remove "_s5cmd"
            method = "s5cmd"
        else:
            base_provider = provider_name
            method = "sdk"

        for size_bytes, size_result in provider_result.size_results.items():
            base = {
                "date": result.timestamp.strftime("%Y-%m-%d"),
                "timestamp": result.timestamp.isoformat(),
                "hostname": machine_info["hostname"],
                "ip_address": machine_info["ip_address"],
                "provider": base_provider,
                "method": method,
                "size_bytes": size_result.size_bytes,
                "size_label": size_result.size_label,
            }
            # Individual upload iterations
            for i, val in enumerate(size_result.upload_throughputs, 1):
                rows.append({**base, "iteration": i, "metric": "upload_mbps", "value": round(val, 3)})
            # Individual download iterations
            for i, val in enumerate(size_result.download_throughputs, 1):
                rows.append({**base, "iteration": i, "metric": "download_mbps", "value": round(val, 3)})
            # Individual latency iterations
            for i, val in enumerate(size_result.latencies, 1):
                rows.append({**base, "iteration": i, "metric": "latency_sec", "value": round(val, 6)})

    if not rows:
        return csv_path

    fieldnames = [
        "date", "timestamp", "hostname", "ip_address", "provider", "method",
        "size_bytes", "size_label", "iteration", "metric", "value"
    ]

    with open(csv_path, "a", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        if not file_exists:
            writer.writeheader()
        writer.writerows(rows)

    return csv_path


def get_excel_path() -> Path:
    """Get the Excel results file path."""
    date_str = datetime.now().strftime("%y%m%d")
    return get_results_dir() / f"{date_str}-benchmarks.xlsx"


def save_results_excel(result: BenchmarkResult) -> Path:
    """Append benchmark results to Excel file (raw data)."""
    excel_path = get_excel_path()
    machine_info = get_machine_info()

    # Collect rows (same format as CSV)
    rows = []
    for provider_name, provider_result in result.provider_results.items():
        if provider_name.endswith("_rclone"):
            base_provider = provider_name[:-7]
            method = "rclone"
        elif provider_name.endswith("_s5cmd"):
            base_provider = provider_name[:-6]
            method = "s5cmd"
        else:
            base_provider = provider_name
            method = "sdk"

        for size_bytes, size_result in provider_result.size_results.items():
            base = {
                "date": result.timestamp.strftime("%Y-%m-%d"),
                "timestamp": result.timestamp.isoformat(),
                "hostname": machine_info["hostname"],
                "ip_address": machine_info["ip_address"],
                "provider": base_provider,
                "method": method,
                "size_bytes": size_result.size_bytes,
                "size_label": size_result.size_label,
            }
            for i, val in enumerate(size_result.upload_throughputs, 1):
                rows.append({**base, "iteration": i, "metric": "upload_mbps", "value": round(val, 3)})
            for i, val in enumerate(size_result.download_throughputs, 1):
                rows.append({**base, "iteration": i, "metric": "download_mbps", "value": round(val, 3)})
            for i, val in enumerate(size_result.latencies, 1):
                rows.append({**base, "iteration": i, "metric": "latency_sec", "value": round(val, 6)})

    if not rows:
        return excel_path

    fieldnames = [
        "date", "timestamp", "hostname", "ip_address", "provider", "method",
        "size_bytes", "size_label", "iteration", "metric", "value"
    ]

    # Load existing or create new workbook
    if excel_path.exists():
        wb = load_workbook(excel_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Benchmarks"
        ws.append(fieldnames)

    # Append rows
    for row in rows:
        ws.append([row[f] for f in fieldnames])

    wb.save(excel_path)
    return excel_path


def save_full_report_excel(
    benchmark_result: BenchmarkResult,
    feature_results: dict = None,
    filename: str = None
) -> Path:
    """Save comprehensive Excel report with multiple sheets."""
    import statistics
    from s3bench.benchmark import get_sizes_for_category

    machine_info = get_machine_info()
    date_str = benchmark_result.timestamp.strftime("%y%m%d")

    if filename:
        excel_path = get_results_dir() / filename
    else:
        excel_path = get_results_dir() / f"{date_str}-s3bench-results.xlsx"

    wb = Workbook()

    # Get all sizes
    all_sizes = []
    for cat in ["small", "medium", "large", "xlarge"]:
        all_sizes.extend(get_sizes_for_category(cat))

    # Group results by base provider and method
    grouped = {}
    for provider_name, provider_result in benchmark_result.provider_results.items():
        base_provider, method = _parse_provider_method(provider_name)
        grouped[(base_provider, method)] = provider_result
    sorted_keys = sorted(grouped.keys(), key=lambda x: (x[0], x[1]))

    # Determine n per size
    size_n = {}
    for pr in benchmark_result.provider_results.values():
        for size in all_sizes:
            sr = pr.size_results.get(size.value)
            if sr and sr.upload_throughputs and size.value not in size_n:
                size_n[size.value] = len(sr.upload_throughputs)

    def size_header(size) -> str:
        n = size_n.get(size.value, "?")
        label = size.name.replace("SMALL_", "").replace("MEDIUM_", "").replace("LARGE_", "").replace("XLARGE_", "")
        return f"{label} (n={n})"

    def calc_std(values):
        if len(values) > 1:
            return statistics.stdev(values)
        return 0.0

    # Sheet 1: Summary info
    ws_info = wb.active
    ws_info.title = "Info"
    ws_info.append(["S3 Benchmark Report"])
    ws_info.append(["Date", benchmark_result.timestamp.strftime("%Y-%m-%d")])
    ws_info.append(["Time", benchmark_result.timestamp.strftime("%H:%M:%S")])
    ws_info.append(["Hostname", machine_info["hostname"]])
    ws_info.append(["IP Address", machine_info["ip_address"]])
    providers_list = list(set(k[0] for k in sorted_keys))
    methods_list = list(set(k[1] for k in sorted_keys))
    ws_info.append(["Providers", ", ".join(sorted(providers_list))])
    ws_info.append(["Methods", ", ".join(sorted(methods_list))])

    # Sheet 2: Upload Throughput (MiB/s) - mean
    ws_upload = wb.create_sheet("Upload Mean")
    ws_upload.append(["Provider", "Method"] + [size_header(s) for s in all_sizes])
    for (provider, method) in sorted_keys:
        pr = grouped[(provider, method)]
        row = [provider, method]
        for size in all_sizes:
            sr = pr.size_results.get(size.value)
            if sr and sr.upload_throughputs:
                row.append(round(statistics.mean(sr.upload_throughputs), 2))
            else:
                row.append("")
        ws_upload.append(row)

    # Sheet 3: Upload Throughput - std
    ws_upload_std = wb.create_sheet("Upload Std")
    ws_upload_std.append(["Provider", "Method"] + [size_header(s) for s in all_sizes])
    for (provider, method) in sorted_keys:
        pr = grouped[(provider, method)]
        row = [provider, method]
        for size in all_sizes:
            sr = pr.size_results.get(size.value)
            if sr and sr.upload_throughputs:
                row.append(round(calc_std(sr.upload_throughputs), 2))
            else:
                row.append("")
        ws_upload_std.append(row)

    # Sheet 4: Download Throughput (MiB/s) - mean
    ws_download = wb.create_sheet("Download Mean")
    ws_download.append(["Provider", "Method"] + [size_header(s) for s in all_sizes])
    for (provider, method) in sorted_keys:
        pr = grouped[(provider, method)]
        row = [provider, method]
        for size in all_sizes:
            sr = pr.size_results.get(size.value)
            if sr and sr.download_throughputs:
                row.append(round(statistics.mean(sr.download_throughputs), 2))
            else:
                row.append("")
        ws_download.append(row)

    # Sheet 5: Download Throughput - std
    ws_download_std = wb.create_sheet("Download Std")
    ws_download_std.append(["Provider", "Method"] + [size_header(s) for s in all_sizes])
    for (provider, method) in sorted_keys:
        pr = grouped[(provider, method)]
        row = [provider, method]
        for size in all_sizes:
            sr = pr.size_results.get(size.value)
            if sr and sr.download_throughputs:
                row.append(round(calc_std(sr.download_throughputs), 2))
            else:
                row.append("")
        ws_download_std.append(row)

    # Sheet 6: Latency (sec) - mean
    ws_latency = wb.create_sheet("Latency Mean")
    ws_latency.append(["Provider", "Method"] + [size_header(s) for s in all_sizes])
    for (provider, method) in sorted_keys:
        pr = grouped[(provider, method)]
        row = [provider, method]
        for size in all_sizes:
            sr = pr.size_results.get(size.value)
            if sr and sr.latencies:
                row.append(round(statistics.mean(sr.latencies), 4))
            else:
                row.append("")
        ws_latency.append(row)

    # Sheet 7: Latency - std
    ws_latency_std = wb.create_sheet("Latency Std")
    ws_latency_std.append(["Provider", "Method"] + [size_header(s) for s in all_sizes])
    for (provider, method) in sorted_keys:
        pr = grouped[(provider, method)]
        row = [provider, method]
        for size in all_sizes:
            sr = pr.size_results.get(size.value)
            if sr and sr.latencies:
                row.append(round(calc_std(sr.latencies), 4))
            else:
                row.append("")
        ws_latency_std.append(row)

    # Sheet 8: Feature Matrix (if provided)
    if feature_results:
        ws_features = wb.create_sheet("Features")
        feature_providers = list(feature_results.keys())
        feature_names = set()
        for pr in feature_results.values():
            feature_names.update(pr.results.keys())
        feature_names = sorted(feature_names)

        ws_features.append(["Feature"] + feature_providers)
        for feature in feature_names:
            row = [feature]
            for provider in feature_providers:
                pr = feature_results.get(provider)
                if pr:
                    result = pr.results.get(feature)
                    if result:
                        if result.status.value == "supported":
                            row.append("Yes")
                        elif result.status.value == "not_supported":
                            row.append("No")
                        elif result.status.value == "not_applicable":
                            row.append("N/A")
                        else:
                            row.append("Error")
                    else:
                        row.append("-")
                else:
                    row.append("-")
            ws_features.append(row)

    # Sheet 9: Raw Data
    ws_raw = wb.create_sheet("Raw Data")
    fieldnames = [
        "date", "timestamp", "hostname", "ip_address", "provider", "method",
        "size_bytes", "size_label", "iteration", "metric", "value"
    ]
    ws_raw.append(fieldnames)

    for provider_name, provider_result in benchmark_result.provider_results.items():
        base_provider, method = _parse_provider_method(provider_name)

        for size_bytes, size_result in provider_result.size_results.items():
            base_row = [
                benchmark_result.timestamp.strftime("%Y-%m-%d"),
                benchmark_result.timestamp.isoformat(),
                machine_info["hostname"],
                machine_info["ip_address"],
                base_provider,
                method,
                size_result.size_bytes,
                size_result.size_label,
            ]
            for i, val in enumerate(size_result.upload_throughputs, 1):
                ws_raw.append(base_row + [i, "upload_mibps", round(val, 3)])
            for i, val in enumerate(size_result.download_throughputs, 1):
                ws_raw.append(base_row + [i, "download_mibps", round(val, 3)])
            for i, val in enumerate(size_result.latencies, 1):
                ws_raw.append(base_row + [i, "latency_sec", round(val, 6)])

    # Sheet 10: Errors
    ws_errors = wb.create_sheet("Errors")
    ws_errors.append(["Provider", "Method", "Error"])
    for provider_name, provider_result in benchmark_result.provider_results.items():
        base_provider, method = _parse_provider_method(provider_name)
        for error in provider_result.errors:
            ws_errors.append([base_provider, method, error])

    wb.save(excel_path)
    return excel_path


app = typer.Typer(help="S3 Provider Benchmark Tool")
provider_app = typer.Typer(help="Manage S3 providers")
app.add_typer(provider_app, name="provider")

console = Console()


@provider_app.command("add")
def provider_add(
    name: Annotated[str, typer.Argument(help="Provider name")],
    bucket: Annotated[str, typer.Option("--bucket", "-b", help="Bucket/container name", prompt=True)],
    access_key: Annotated[str, typer.Option("--access-key", "-a", help="Access key (or Azure storage account name)", prompt=True)],
    secret_key: Annotated[
        str, typer.Option("--secret-key", "-s", help="Secret key", prompt=True, hide_input=True)
    ],
    endpoint_url: Annotated[
        Optional[str], typer.Option("--endpoint", "-e", help="Custom endpoint URL")
    ] = None,
    iam_endpoint: Annotated[
        Optional[str], typer.Option("--iam-endpoint", help="IAM API endpoint (for credential vending)")
    ] = None,
    region: Annotated[Optional[str], typer.Option("--region", "-r", help="Region")] = None,
    provider_type: Annotated[
        ProviderType, typer.Option("--type", "-t", help="Provider type (s3 or azure)")
    ] = ProviderType.S3,
) -> None:
    """Add a new storage provider."""
    cfg.add_provider(
        name=name,
        bucket=bucket,
        access_key=access_key,
        secret_key=secret_key,
        endpoint_url=endpoint_url,
        iam_endpoint=iam_endpoint,
        region=region,
        provider_type=provider_type,
    )
    console.print(f"Provider [bold green]{name}[/] ({provider_type.value}) added successfully.")


@provider_app.command("list")
def provider_list() -> None:
    """List all configured providers."""
    config = cfg.load_config()

    if not config.providers:
        console.print("No providers configured. Use [bold]s3bench provider add[/] to add one.")
        return

    table = Table(title="Configured Providers")
    table.add_column("Name", style="cyan")
    table.add_column("Type")
    table.add_column("Endpoint")
    table.add_column("Region")
    table.add_column("Bucket", style="green")

    for name, provider in config.providers.items():
        table.add_row(
            name,
            provider.provider_type.value,
            provider.endpoint_url or "(default)",
            provider.region or "-",
            provider.bucket,
        )

    console.print(table)


@provider_app.command("remove")
def provider_remove(
    name: Annotated[str, typer.Argument(help="Provider name to remove")],
) -> None:
    """Remove a provider."""
    if cfg.remove_provider(name):
        console.print(f"Provider [bold red]{name}[/] removed.")
    else:
        console.print(f"Provider [bold]{name}[/] not found.", style="red")
        raise typer.Exit(1)


@provider_app.command("test")
def provider_test(
    name: Annotated[str, typer.Argument(help="Provider name to test")],
) -> None:
    """Test connection to a provider."""
    provider_config = cfg.get_provider(name)

    if not provider_config:
        console.print(f"Provider [bold]{name}[/] not found.", style="red")
        raise typer.Exit(1)

    provider = create_provider(name, provider_config)

    with console.status(f"Testing connection to {name}..."):
        if provider.test_connection():
            console.print(f"[bold green]Success![/] Connected to {name}")
        else:
            console.print(f"[bold red]Failed![/] Could not connect to {name}")
            raise typer.Exit(1)


@provider_app.command("cleanup")
def provider_cleanup(
    name: Annotated[str, typer.Argument(help="Provider name to clean up")],
) -> None:
    """Remove leftover benchmark files from a provider."""
    provider_config = cfg.get_provider(name)

    if not provider_config:
        console.print(f"Provider [bold]{name}[/] not found.", style="red")
        raise typer.Exit(1)

    provider = create_provider(name, provider_config)

    with console.status(f"Cleaning up benchmark files from {name}..."):
        count = provider.cleanup_benchmark_objects()

    if count > 0:
        console.print(f"Deleted [bold]{count}[/] benchmark files from {name}")
    else:
        console.print(f"No benchmark files found in {name}")


@app.command("run")
def run(
    providers: Annotated[
        Optional[list[str]],
        typer.Option("--provider", "-p", help="Provider(s) to benchmark"),
    ] = None,
    all_providers: Annotated[
        bool, typer.Option("--all", "-a", help="Run against all providers, methods, and sizes")
    ] = False,
    sizes: Annotated[
        Optional[str], typer.Option("--sizes", "-s", help="Size categories: small,medium,large,xlarge")
    ] = None,
    small_iter: Annotated[
        int, typer.Option("--small-iter", help="Iterations for small files")
    ] = DEFAULT_ITERATIONS["small"],
    medium_iter: Annotated[
        int, typer.Option("--medium-iter", help="Iterations for medium files")
    ] = DEFAULT_ITERATIONS["medium"],
    large_iter: Annotated[
        int, typer.Option("--large-iter", help="Iterations for large files")
    ] = DEFAULT_ITERATIONS["large"],
    xlarge_iter: Annotated[
        int, typer.Option("--xlarge-iter", help="Iterations for xlarge files (1GB, 4GB)")
    ] = DEFAULT_ITERATIONS["xlarge"],
    method: Annotated[
        Optional[str], typer.Option("--method", "-m", help="Benchmark method: sdk, rclone, s5cmd, or all")
    ] = None,
) -> None:
    """Run benchmarks against providers."""
    config = cfg.load_config()

    if not config.providers:
        console.print("No providers configured. Use [bold]s3bench provider add[/] first.")
        raise typer.Exit(1)

    # Default sizes: include xlarge if --all, otherwise exclude
    if sizes is None:
        sizes = "small,medium,large,xlarge" if all_providers else "small,medium,large"

    # Default method: "all" if --all flag, otherwise "sdk"
    if method is None:
        method = "all" if all_providers else "sdk"

    # Validate method (support comma-separated list)
    methods = [m.strip() for m in method.split(",")]
    valid_methods = {"sdk", "rclone", "s5cmd", "all"}
    for m in methods:
        if m not in valid_methods:
            console.print(f"Invalid method: {m}. Use: sdk, rclone, s5cmd, all, or comma-separated", style="red")
            raise typer.Exit(1)
    if "all" in methods:
        methods = ["sdk", "rclone", "s5cmd"]

    # Check rclone if needed
    if "rclone" in methods and not check_rclone_installed():
        console.print("rclone is not installed. Install from https://rclone.org/install/", style="red")
        raise typer.Exit(1)

    # Check s5cmd if needed
    if "s5cmd" in methods and not check_s5cmd_installed():
        console.print("s5cmd is not installed. Install from https://github.com/peak/s5cmd", style="red")
        raise typer.Exit(1)

    # Determine which providers to run
    if all_providers:
        provider_names = list(config.providers.keys())
    elif providers:
        provider_names = providers
    else:
        console.print("Specify providers with --provider or use --all")
        raise typer.Exit(1)

    # Validate providers exist and create instances
    storage_providers = []
    rclone_providers = []
    s5cmd_providers = []
    for name in provider_names:
        provider_config = config.providers.get(name)
        if not provider_config:
            console.print(f"Provider [bold]{name}[/] not found.", style="red")
            raise typer.Exit(1)
        if "sdk" in methods:
            storage_providers.append(create_provider(name, provider_config))
        if "rclone" in methods:
            rclone_providers.append(create_rclone_provider(f"{name}_rclone", provider_config))
        if "s5cmd" in methods:
            # s5cmd only supports S3-compatible providers
            if provider_config.provider_type != ProviderType.AZURE:
                s5cmd_providers.append(create_s5cmd_provider(f"{name}_s5cmd", provider_config))
            else:
                console.print(f"[yellow]Skipping s5cmd for {name} (Azure not supported by s5cmd)[/]")

    # Parse size categories
    categories = [s.strip() for s in sizes.split(",")]
    valid_categories = {"small", "medium", "large", "xlarge"}
    for cat in categories:
        if cat not in valid_categories:
            console.print(f"Invalid size category: {cat}", style="red")
            raise typer.Exit(1)

    iterations = {
        "small": small_iter,
        "medium": medium_iter,
        "large": large_iter,
        "xlarge": xlarge_iter,
    }

    # Combine all providers for benchmarking
    all_benchmark_providers = storage_providers + rclone_providers + s5cmd_providers

    method_label = ", ".join(m.upper() if m == "sdk" else m for m in methods)
    console.print(f"Running benchmarks for: [bold]{', '.join(provider_names)}[/]")
    console.print(f"Method: [bold]{method_label}[/]")
    console.print(f"Size categories: [bold]{', '.join(categories)}[/]")

    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        TaskProgressColumn(),
        console=console,
    ) as progress:
        task = progress.add_task("Starting...", total=100)

        def update_progress(msg: str, current: int, total: int) -> None:
            progress.update(task, completed=(current / total) * 100, description=msg)

        result = run_benchmark(
            providers=all_benchmark_providers,
            categories=categories,
            iterations=iterations,
            progress_callback=update_progress,
        )

    # Cleanup temp files
    for rp in rclone_providers:
        rp.cleanup()
    for sp in s5cmd_providers:
        sp.cleanup()

    # Save and display results
    results_file = save_results(result)
    csv_file = save_results_csv(result)
    excel_file = save_results_excel(result)
    console.print(f"\nResults saved to: [dim]{results_file}[/]")
    console.print(f"CSV appended to: [dim]{csv_file}[/]")
    console.print(f"Excel appended to: [dim]{excel_file}[/]\n")
    _display_results(result, categories)


def _display_results(result, categories: list[str]) -> None:
    """Display benchmark results as tables with provider/method as rows and sizes as columns."""
    import statistics
    from s3bench.benchmark import get_sizes_for_category

    # Get all sizes tested
    all_sizes = []
    for cat in categories:
        all_sizes.extend(get_sizes_for_category(cat))

    # Group results by base provider and method
    grouped = {}  # {(provider, method): provider_result}
    for provider_name, provider_result in result.provider_results.items():
        base_provider, method = _parse_provider_method(provider_name)
        grouped[(base_provider, method)] = provider_result

    # Sort by provider then method
    sorted_keys = sorted(grouped.keys(), key=lambda x: (x[0], x[1]))

    # Determine n (iterations) per size category from first result with data
    size_n = {}
    for pr in result.provider_results.values():
        for size in all_sizes:
            sr = pr.size_results.get(size.value)
            if sr and sr.upload_throughputs and size.value not in size_n:
                size_n[size.value] = len(sr.upload_throughputs)

    # Build size label without prefix
    def size_label(size) -> str:
        return size.name.replace("SMALL_", "").replace("MEDIUM_", "").replace("LARGE_", "").replace("XLARGE_", "")

    def calc_mean(values):
        return statistics.mean(values) if values else 0.0

    def calc_std(values):
        return statistics.stdev(values) if len(values) > 1 else 0.0

    # Upload throughput table - separate mean and std columns
    upload_table = Table(title="Upload Throughput (MiB/s)")
    upload_table.add_column("Provider", style="cyan")
    upload_table.add_column("Method", style="dim")
    for size in all_sizes:
        n = size_n.get(size.value, "?")
        upload_table.add_column(f"{size_label(size)} mean", justify="right")
        upload_table.add_column(f"std (n={n})", justify="right", style="dim")

    for (provider, method) in sorted_keys:
        pr = grouped[(provider, method)]
        row = [provider, method]
        for size in all_sizes:
            sr = pr.size_results.get(size.value)
            if sr and sr.upload_throughputs:
                row.append(f"{calc_mean(sr.upload_throughputs):.2f}")
                row.append(f"{calc_std(sr.upload_throughputs):.2f}")
            else:
                row.extend(["-", "-"])
        upload_table.add_row(*row)

    console.print(upload_table)
    console.print()

    # Download throughput table
    download_table = Table(title="Download Throughput (MiB/s)")
    download_table.add_column("Provider", style="cyan")
    download_table.add_column("Method", style="dim")
    for size in all_sizes:
        n = size_n.get(size.value, "?")
        download_table.add_column(f"{size_label(size)} mean", justify="right")
        download_table.add_column(f"std (n={n})", justify="right", style="dim")

    for (provider, method) in sorted_keys:
        pr = grouped[(provider, method)]
        row = [provider, method]
        for size in all_sizes:
            sr = pr.size_results.get(size.value)
            if sr and sr.download_throughputs:
                row.append(f"{calc_mean(sr.download_throughputs):.2f}")
                row.append(f"{calc_std(sr.download_throughputs):.2f}")
            else:
                row.extend(["-", "-"])
        download_table.add_row(*row)

    console.print(download_table)
    console.print()

    # Latency table
    latency_table = Table(title="Latency / TTFB (seconds)")
    latency_table.add_column("Provider", style="cyan")
    latency_table.add_column("Method", style="dim")
    for size in all_sizes:
        n = size_n.get(size.value, "?")
        latency_table.add_column(f"{size_label(size)} mean", justify="right")
        latency_table.add_column(f"std (n={n})", justify="right", style="dim")

    for (provider, method) in sorted_keys:
        pr = grouped[(provider, method)]
        row = [provider, method]
        for size in all_sizes:
            sr = pr.size_results.get(size.value)
            if sr and sr.latencies:
                row.append(f"{calc_mean(sr.latencies):.3f}")
                row.append(f"{calc_std(sr.latencies):.3f}")
            else:
                row.extend(["-", "-"])
        latency_table.add_row(*row)

    console.print(latency_table)

    # Show errors if any
    for provider_name, provider_result in result.provider_results.items():
        if provider_result.errors:
            console.print(f"\n[bold red]Errors for {provider_name}:[/]")
            for error in provider_result.errors:
                console.print(f"  - {error}")


def _load_benchmark_result(data: dict) -> BenchmarkResult:
    """Load a BenchmarkResult from JSON data."""
    from s3bench.benchmark import SizeResult, ProviderResult

    timestamp = datetime.fromisoformat(data["timestamp"])
    result = BenchmarkResult(timestamp=timestamp)

    for provider_name, provider_data in data.get("providers", {}).items():
        pr = ProviderResult(provider_name=provider_name)
        pr.errors = provider_data.get("errors", [])

        # Handle both "size_results" (current) and "sizes" (legacy) keys
        size_results = provider_data.get("size_results", provider_data.get("sizes", {}))
        for size_str, size_data in size_results.items():
            size_bytes = int(size_str)
            sr = SizeResult(size_bytes=size_bytes)
            sr.upload_throughputs = size_data.get("upload_throughputs", [])
            sr.download_throughputs = size_data.get("download_throughputs", [])
            sr.latencies = size_data.get("latencies", [])
            pr.size_results[size_bytes] = sr

        result.provider_results[provider_name] = pr

    return result


def _infer_categories_from_result(result: BenchmarkResult) -> list[str]:
    """Infer size categories from a benchmark result."""
    from s3bench.benchmark import FileSize

    # Collect all size bytes from the result
    all_sizes = set()
    for pr in result.provider_results.values():
        all_sizes.update(pr.size_results.keys())

    # Map to categories
    categories = set()
    for size in FileSize:
        if size.value in all_sizes:
            if size.name.startswith("SMALL_"):
                categories.add("small")
            elif size.name.startswith("MEDIUM_"):
                categories.add("medium")
            elif size.name.startswith("LARGE_"):
                categories.add("large")
            elif size.name.startswith("XLARGE_"):
                categories.add("xlarge")

    # Return in order
    order = ["small", "medium", "large", "xlarge"]
    return [c for c in order if c in categories]


@app.command("results")
def results(
    last: Annotated[int, typer.Option("--last", "-n", help="Show last N results")] = 5,
    show: Annotated[bool, typer.Option("--show", "-s", help="Display benchmark tables from last result")] = False,
) -> None:
    """List past benchmark results."""
    results_dir = get_results_dir()
    files = sorted(results_dir.glob("benchmark_*.json"), reverse=True)

    if not files:
        console.print("No benchmark results found.")
        return

    if show:
        # Load and display the last result
        with open(files[0]) as fp:
            data = json.load(fp)

        # Reconstruct BenchmarkResult from JSON
        result = _load_benchmark_result(data)

        # Determine categories from the sizes in the data
        categories = _infer_categories_from_result(result)

        console.print(f"[dim]Results from: {files[0].name}[/]\n")
        _display_results(result, categories)
        return

    table = Table(title="Benchmark Results")
    table.add_column("Date", style="cyan")
    table.add_column("Providers")
    table.add_column("File")

    for f in files[:last]:
        with open(f) as fp:
            data = json.load(fp)
        providers = ", ".join(data.get("providers", {}).keys())
        timestamp = data.get("timestamp", "unknown")
        table.add_row(timestamp[:19], providers, f.name)

    console.print(table)
    console.print(f"\nResults stored in: [dim]{results_dir}[/]")
    console.print("[dim]Use --show to display the last benchmark tables[/]")


def _save_features_excel(all_results: dict) -> Path:
    """Save feature test results to Excel."""
    from datetime import datetime

    date_str = datetime.now().strftime("%y%m%d")
    excel_path = get_results_dir() / f"{date_str}-features.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Features"

    # Get all feature names
    feature_names = set()
    for provider_results in all_results.values():
        feature_names.update(provider_results.results.keys())
    feature_names = sorted(feature_names)

    provider_names = list(all_results.keys())

    # Header row
    ws.append(["Feature"] + provider_names + ["Description"])

    # Data rows
    for feature in feature_names:
        row = [feature]
        for provider in provider_names:
            pr = all_results.get(provider)
            if pr:
                result = pr.results.get(feature)
                if result:
                    if result.status.value == "supported":
                        row.append("Yes")
                    elif result.status.value == "not_supported":
                        row.append("No")
                    elif result.status.value == "not_applicable":
                        row.append("N/A")
                    else:
                        row.append("Error")
                else:
                    row.append("-")
            else:
                row.append("-")
        # Add description as last column
        row.append(FEATURE_DESCRIPTIONS.get(feature, ""))
        ws.append(row)

    wb.save(excel_path)
    return excel_path


@app.command("features")
def features(
    providers: Annotated[
        Optional[list[str]],
        typer.Option("--provider", "-p", help="Provider(s) to test"),
    ] = None,
    all_providers: Annotated[
        bool, typer.Option("--all", "-a", help="Test all providers")
    ] = False,
    feature_list: Annotated[
        Optional[str], typer.Option("--features", "-f", help="Comma-separated features to test")
    ] = None,
    output_json: Annotated[
        bool, typer.Option("--json", help="Output as JSON")
    ] = False,
) -> None:
    """Test S3 API feature support across providers."""
    config = cfg.load_config()

    if not config.providers:
        console.print("No providers configured. Use [bold]s3bench provider add[/] first.")
        raise typer.Exit(1)

    # Determine which providers to test
    if all_providers:
        provider_names = list(config.providers.keys())
    elif providers:
        provider_names = providers
    else:
        console.print("Specify providers with --provider or use --all")
        raise typer.Exit(1)

    # Parse features
    features_to_test = None
    if feature_list:
        features_to_test = [f.strip() for f in feature_list.split(",")]
        for f in features_to_test:
            if f not in FEATURE_TESTS:
                console.print(f"Unknown feature: {f}", style="red")
                console.print(f"Available: {', '.join(FEATURE_TESTS.keys())}")
                raise typer.Exit(1)

    # Validate providers and create instances
    storage_providers = []
    for name in provider_names:
        provider_config = config.providers.get(name)
        if not provider_config:
            console.print(f"Provider [bold]{name}[/] not found.", style="red")
            raise typer.Exit(1)
        storage_providers.append(create_provider(name, provider_config))

    console.print(f"Testing features for: [bold]{', '.join(provider_names)}[/]")

    # Run feature tests
    all_results = {}
    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        console=console,
    ) as progress:
        for provider in storage_providers:
            task = progress.add_task(f"Testing {provider.name}...", total=None)

            def update_progress(feature: str, current: int, total: int) -> None:
                progress.update(task, description=f"{provider.name}: {feature}")

            results = run_feature_tests(provider, features_to_test, update_progress)
            all_results[provider.name] = results
            progress.update(task, description=f"{provider.name}: Done")

    # Save to Excel
    excel_path = _save_features_excel(all_results)

    # Output results
    if output_json:
        json_output = {}
        for provider_name, provider_results in all_results.items():
            json_output[provider_name] = {
                name: {
                    "status": result.status.value,
                    "message": result.message,
                }
                for name, result in provider_results.results.items()
            }
        console.print(json.dumps(json_output, indent=2))
    else:
        _display_feature_results(all_results)

    console.print(f"\nResults saved to: [cyan]{excel_path}[/]")


def _display_feature_results(all_results: dict) -> None:
    """Display feature test results as a table."""
    # Get all feature names from results
    feature_names = set()
    for provider_results in all_results.values():
        feature_names.update(provider_results.results.keys())
    feature_names = sorted(feature_names)

    # Create table
    table = Table(title="Feature Support Matrix")
    table.add_column("Feature", style="cyan")

    provider_names = list(all_results.keys())
    for name in provider_names:
        table.add_column(name, justify="center")

    table.add_column("Description", style="dim")

    # Add rows
    for feature_name in feature_names:
        row = [feature_name]
        for provider_name in provider_names:
            provider_results = all_results[provider_name]
            result = provider_results.results.get(feature_name)
            if result:
                if result.status == FeatureStatus.SUPPORTED:
                    cell = "[green]Yes[/]"
                elif result.status == FeatureStatus.NOT_SUPPORTED:
                    cell = "[red]No[/]"
                elif result.status == FeatureStatus.NOT_APPLICABLE:
                    cell = "[dim]N/A[/]"
                else:
                    cell = f"[yellow]Error[/]"
            else:
                cell = "-"
            row.append(cell)
        # Add description as last column
        description = FEATURE_DESCRIPTIONS.get(feature_name, "")
        row.append(description)
        table.add_row(*row)

    console.print()
    console.print(table)

    # Show any error messages
    has_errors = False
    for provider_name, provider_results in all_results.items():
        for feature_name, result in provider_results.results.items():
            if result.status == FeatureStatus.ERROR and result.message:
                if not has_errors:
                    console.print("\n[bold yellow]Errors:[/]")
                    has_errors = True
                console.print(f"  {provider_name}/{feature_name}: {result.message}")


@app.command("report")
def report(
    providers: Annotated[
        Optional[list[str]],
        typer.Option("--provider", "-p", help="Provider(s) to test"),
    ] = None,
    all_providers: Annotated[
        bool, typer.Option("--all", "-a", help="Run against all providers, methods, and sizes")
    ] = False,
    sizes: Annotated[
        Optional[str], typer.Option("--sizes", "-s", help="Size categories: small,medium,large,xlarge")
    ] = None,
    small_iter: Annotated[
        int, typer.Option("--small-iter", help="Iterations for small files")
    ] = DEFAULT_ITERATIONS["small"],
    medium_iter: Annotated[
        int, typer.Option("--medium-iter", help="Iterations for medium files")
    ] = DEFAULT_ITERATIONS["medium"],
    large_iter: Annotated[
        int, typer.Option("--large-iter", help="Iterations for large files")
    ] = DEFAULT_ITERATIONS["large"],
    xlarge_iter: Annotated[
        int, typer.Option("--xlarge-iter", help="Iterations for xlarge files (1GB, 4GB)")
    ] = DEFAULT_ITERATIONS["xlarge"],
    method: Annotated[
        Optional[str], typer.Option("--method", "-m", help="Benchmark method: sdk, rclone, s5cmd, or all")
    ] = None,
    output: Annotated[
        Optional[str], typer.Option("--output", "-o", help="Output filename (default: YYMMDD-s3bench-results.xlsx)")
    ] = None,
) -> None:
    """Run full report: benchmarks + features -> Excel."""
    config = cfg.load_config()

    if not config.providers:
        console.print("No providers configured. Use [bold]s3bench provider add[/] first.")
        raise typer.Exit(1)

    # Default sizes: include xlarge if --all, otherwise exclude
    if sizes is None:
        sizes = "small,medium,large,xlarge" if all_providers else "small,medium,large"

    # Default method: "all" if --all flag, otherwise "sdk"
    if method is None:
        method = "all" if all_providers else "sdk"

    # Validate method (support comma-separated list)
    methods = [m.strip() for m in method.split(",")]
    valid_methods = {"sdk", "rclone", "s5cmd", "all"}
    for m in methods:
        if m not in valid_methods:
            console.print(f"Invalid method: {m}. Use: sdk, rclone, s5cmd, all, or comma-separated", style="red")
            raise typer.Exit(1)
    if "all" in methods:
        methods = ["sdk", "rclone", "s5cmd"]

    # Check rclone if needed
    if "rclone" in methods and not check_rclone_installed():
        console.print("rclone is not installed. Install from https://rclone.org/install/", style="red")
        raise typer.Exit(1)

    # Check s5cmd if needed
    if "s5cmd" in methods and not check_s5cmd_installed():
        console.print("s5cmd is not installed. Install from https://github.com/peak/s5cmd", style="red")
        raise typer.Exit(1)

    # Determine which providers to run
    if all_providers:
        provider_names = list(config.providers.keys())
    elif providers:
        provider_names = providers
    else:
        console.print("Specify providers with --provider or use --all")
        raise typer.Exit(1)

    # Validate providers exist and create instances
    storage_providers = []
    rclone_providers = []
    s5cmd_providers = []
    for name in provider_names:
        provider_config = config.providers.get(name)
        if not provider_config:
            console.print(f"Provider [bold]{name}[/] not found.", style="red")
            raise typer.Exit(1)
        if "sdk" in methods:
            storage_providers.append(create_provider(name, provider_config))
        if "rclone" in methods:
            rclone_providers.append(create_rclone_provider(f"{name}_rclone", provider_config))
        if "s5cmd" in methods:
            # s5cmd only supports S3-compatible providers
            if provider_config.provider_type != ProviderType.AZURE:
                s5cmd_providers.append(create_s5cmd_provider(f"{name}_s5cmd", provider_config))
            else:
                console.print(f"[yellow]Skipping s5cmd for {name} (Azure not supported by s5cmd)[/]")

    # Parse size categories
    categories = [s.strip() for s in sizes.split(",")]
    valid_categories = {"small", "medium", "large", "xlarge"}
    for cat in categories:
        if cat not in valid_categories:
            console.print(f"Invalid size category: {cat}", style="red")
            raise typer.Exit(1)

    iterations = {
        "small": small_iter,
        "medium": medium_iter,
        "large": large_iter,
        "xlarge": xlarge_iter,
    }

    all_benchmark_providers = storage_providers + rclone_providers + s5cmd_providers

    method_label = ", ".join(m.upper() if m == "sdk" else m for m in methods)
    console.print(f"\n[bold]S3 Benchmark Report[/]")
    console.print(f"Providers: [bold]{', '.join(provider_names)}[/]")
    console.print(f"Method: [bold]{method_label}[/]")
    console.print(f"Size categories: [bold]{', '.join(categories)}[/]\n")

    # Run benchmarks
    console.print("[bold]Running benchmarks...[/]")
    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        TaskProgressColumn(),
        console=console,
    ) as progress:
        task = progress.add_task("Starting...", total=100)

        def update_progress(msg: str, current: int, total: int) -> None:
            progress.update(task, completed=(current / total) * 100, description=msg)

        benchmark_result = run_benchmark(
            providers=all_benchmark_providers,
            categories=categories,
            iterations=iterations,
            progress_callback=update_progress,
        )

    # Cleanup temp files
    for rp in rclone_providers:
        rp.cleanup()
    for sp in s5cmd_providers:
        sp.cleanup()

    # Run feature tests (SDK providers only)
    console.print("\n[bold]Running feature tests...[/]")
    feature_results = {}
    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        console=console,
    ) as progress:
        for provider in storage_providers:
            task = progress.add_task(f"Testing {provider.name}...", total=None)

            def update_progress(feature: str, current: int, total: int) -> None:
                progress.update(task, description=f"{provider.name}: {feature}")

            results = run_feature_tests(provider, None, update_progress)
            feature_results[provider.name] = results
            progress.update(task, description=f"{provider.name}: Done")

    # Save all results
    results_file = save_results(benchmark_result)
    csv_file = save_results_csv(benchmark_result)
    excel_file = save_results_excel(benchmark_result)
    report_file = save_full_report_excel(benchmark_result, feature_results, output)

    console.print(f"\n[bold green]Report generated![/]")
    console.print(f"  JSON: [dim]{results_file}[/]")
    console.print(f"  CSV: [dim]{csv_file}[/]")
    console.print(f"  Excel (raw): [dim]{excel_file}[/]")
    console.print(f"  Excel (report): [bold]{report_file}[/]\n")

    # Display results
    _display_results(benchmark_result, categories)
    _display_feature_results(feature_results)


if __name__ == "__main__":
    app()
