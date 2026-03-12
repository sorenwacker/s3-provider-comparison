"""Tests for the historical report functionality in CLI."""

import csv
import statistics
from datetime import datetime
from pathlib import Path
from tempfile import TemporaryDirectory

import pytest
from openpyxl import load_workbook

from s3bench.cli import (
    get_csv_path,
    load_historical_data,
    save_historical_report_excel,
)


@pytest.fixture
def temp_results_dir(monkeypatch):
    """Create a temporary results directory."""
    with TemporaryDirectory() as tmpdir:
        tmpdir_path = Path(tmpdir)
        monkeypatch.setattr("s3bench.cli.get_results_dir", lambda: tmpdir_path)
        yield tmpdir_path


@pytest.fixture
def sample_csv_data():
    """Sample CSV data for testing."""
    return [
        # Day 1: aws sdk
        {"date": "2026-03-01", "timestamp": "2026-03-01T10:00:00", "hostname": "host1",
         "ip_address": "192.168.1.1", "provider": "aws", "method": "sdk",
         "size_bytes": "1024", "size_label": "1KB", "iteration": "1",
         "metric": "upload_mbps", "value": "10.5"},
        {"date": "2026-03-01", "timestamp": "2026-03-01T10:00:00", "hostname": "host1",
         "ip_address": "192.168.1.1", "provider": "aws", "method": "sdk",
         "size_bytes": "1024", "size_label": "1KB", "iteration": "2",
         "metric": "upload_mbps", "value": "11.5"},
        {"date": "2026-03-01", "timestamp": "2026-03-01T10:00:00", "hostname": "host1",
         "ip_address": "192.168.1.1", "provider": "aws", "method": "sdk",
         "size_bytes": "1024", "size_label": "1KB", "iteration": "1",
         "metric": "download_mbps", "value": "20.0"},
        {"date": "2026-03-01", "timestamp": "2026-03-01T10:00:00", "hostname": "host1",
         "ip_address": "192.168.1.1", "provider": "aws", "method": "sdk",
         "size_bytes": "1024", "size_label": "1KB", "iteration": "2",
         "metric": "download_mbps", "value": "22.0"},
        {"date": "2026-03-01", "timestamp": "2026-03-01T10:00:00", "hostname": "host1",
         "ip_address": "192.168.1.1", "provider": "aws", "method": "sdk",
         "size_bytes": "1024", "size_label": "1KB", "iteration": "1",
         "metric": "latency_sec", "value": "0.05"},
        {"date": "2026-03-01", "timestamp": "2026-03-01T10:00:00", "hostname": "host1",
         "ip_address": "192.168.1.1", "provider": "aws", "method": "sdk",
         "size_bytes": "1024", "size_label": "1KB", "iteration": "2",
         "metric": "latency_sec", "value": "0.06"},
        # Day 1: wasabi sdk
        {"date": "2026-03-01", "timestamp": "2026-03-01T10:00:00", "hostname": "host1",
         "ip_address": "192.168.1.1", "provider": "wasabi", "method": "sdk",
         "size_bytes": "1024", "size_label": "1KB", "iteration": "1",
         "metric": "upload_mbps", "value": "8.0"},
        {"date": "2026-03-01", "timestamp": "2026-03-01T10:00:00", "hostname": "host1",
         "ip_address": "192.168.1.1", "provider": "wasabi", "method": "sdk",
         "size_bytes": "1024", "size_label": "1KB", "iteration": "1",
         "metric": "download_mbps", "value": "15.0"},
        {"date": "2026-03-01", "timestamp": "2026-03-01T10:00:00", "hostname": "host1",
         "ip_address": "192.168.1.1", "provider": "wasabi", "method": "sdk",
         "size_bytes": "1024", "size_label": "1KB", "iteration": "1",
         "metric": "latency_sec", "value": "0.08"},
        # Day 2: aws sdk (different values)
        {"date": "2026-03-05", "timestamp": "2026-03-05T14:00:00", "hostname": "host1",
         "ip_address": "192.168.1.1", "provider": "aws", "method": "sdk",
         "size_bytes": "1024", "size_label": "1KB", "iteration": "1",
         "metric": "upload_mbps", "value": "12.0"},
        {"date": "2026-03-05", "timestamp": "2026-03-05T14:00:00", "hostname": "host1",
         "ip_address": "192.168.1.1", "provider": "aws", "method": "sdk",
         "size_bytes": "1024", "size_label": "1KB", "iteration": "1",
         "metric": "download_mbps", "value": "25.0"},
        {"date": "2026-03-05", "timestamp": "2026-03-05T14:00:00", "hostname": "host1",
         "ip_address": "192.168.1.1", "provider": "aws", "method": "sdk",
         "size_bytes": "1024", "size_label": "1KB", "iteration": "1",
         "metric": "latency_sec", "value": "0.04"},
        # Day 2: aws with 10KB size
        {"date": "2026-03-05", "timestamp": "2026-03-05T14:00:00", "hostname": "host1",
         "ip_address": "192.168.1.1", "provider": "aws", "method": "sdk",
         "size_bytes": "10240", "size_label": "10KB", "iteration": "1",
         "metric": "upload_mbps", "value": "15.0"},
        {"date": "2026-03-05", "timestamp": "2026-03-05T14:00:00", "hostname": "host1",
         "ip_address": "192.168.1.1", "provider": "aws", "method": "sdk",
         "size_bytes": "10240", "size_label": "10KB", "iteration": "1",
         "metric": "download_mbps", "value": "30.0"},
        {"date": "2026-03-05", "timestamp": "2026-03-05T14:00:00", "hostname": "host1",
         "ip_address": "192.168.1.1", "provider": "aws", "method": "sdk",
         "size_bytes": "10240", "size_label": "10KB", "iteration": "1",
         "metric": "latency_sec", "value": "0.03"},
    ]


@pytest.fixture
def csv_file(temp_results_dir, sample_csv_data):
    """Create a sample CSV file with benchmark data."""
    csv_path = temp_results_dir / "benchmarks.csv"
    fieldnames = [
        "date", "timestamp", "hostname", "ip_address", "provider", "method",
        "size_bytes", "size_label", "iteration", "metric", "value"
    ]
    with open(csv_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(sample_csv_data)
    return csv_path


class TestLoadHistoricalData:
    """Tests for load_historical_data function."""

    def test_loads_all_data(self, csv_file, temp_results_dir):
        """Test loading all historical data without filters."""
        data = load_historical_data()

        # Should have data for both dates
        assert len(data["dates"]) == 2
        assert "2026-03-01" in data["dates"]
        assert "2026-03-05" in data["dates"]

    def test_groups_by_date_provider_method(self, csv_file, temp_results_dir):
        """Test data is grouped correctly."""
        data = load_historical_data()

        # Check structure: grouped[date][provider][method][size_bytes][metric] = mean_value
        assert "2026-03-01" in data["grouped"]
        assert "aws" in data["grouped"]["2026-03-01"]
        assert "sdk" in data["grouped"]["2026-03-01"]["aws"]

    def test_calculates_mean_of_iterations(self, csv_file, temp_results_dir):
        """Test that mean is calculated across iterations."""
        data = load_historical_data()

        # Day 1, aws, sdk, 1KB: upload was 10.5 and 11.5, mean should be 11.0
        upload_mean = data["grouped"]["2026-03-01"]["aws"]["sdk"]["1024"]["upload_mbps"]
        assert upload_mean == 11.0

    def test_filter_by_from_date(self, csv_file, temp_results_dir):
        """Test filtering by start date."""
        data = load_historical_data(from_date="2026-03-05")

        # Should only have data from 2026-03-05
        assert len(data["dates"]) == 1
        assert "2026-03-05" in data["dates"]
        assert "2026-03-01" not in data["dates"]

    def test_filter_by_to_date(self, csv_file, temp_results_dir):
        """Test filtering by end date."""
        data = load_historical_data(to_date="2026-03-01")

        # Should only have data up to 2026-03-01
        assert len(data["dates"]) == 1
        assert "2026-03-01" in data["dates"]
        assert "2026-03-05" not in data["dates"]

    def test_filter_by_date_range(self, csv_file, temp_results_dir):
        """Test filtering by date range."""
        data = load_historical_data(from_date="2026-03-01", to_date="2026-03-05")

        # Should have both dates
        assert len(data["dates"]) == 2

    def test_filter_by_providers(self, csv_file, temp_results_dir):
        """Test filtering by provider list."""
        data = load_historical_data(providers=["aws"])

        # Should only have aws data
        assert "aws" in data["providers"]
        assert "wasabi" not in data["providers"]

    def test_returns_empty_for_no_data(self, temp_results_dir):
        """Test returns empty structure when no CSV exists."""
        data = load_historical_data()

        assert data["dates"] == []
        assert data["grouped"] == {}
        assert data["providers"] == set()

    def test_extracts_all_sizes(self, csv_file, temp_results_dir):
        """Test that all size_bytes are extracted."""
        data = load_historical_data()

        # Should have both 1024 and 10240 sizes
        assert "1024" in data["sizes"] or 1024 in data["sizes"]
        assert "10240" in data["sizes"] or 10240 in data["sizes"]

    def test_extracts_size_labels(self, csv_file, temp_results_dir):
        """Test that size labels are mapped correctly."""
        data = load_historical_data()

        assert data["size_labels"]["1024"] == "1KB"
        assert data["size_labels"]["10240"] == "10KB"


class TestSaveHistoricalReportExcel:
    """Tests for save_historical_report_excel function."""

    def test_creates_excel_file(self, csv_file, temp_results_dir):
        """Test that Excel file is created."""
        data = load_historical_data()
        excel_path = save_historical_report_excel(data)

        assert excel_path.exists()
        assert excel_path.suffix == ".xlsx"

    def test_has_summary_sheet(self, csv_file, temp_results_dir):
        """Test that Summary sheet exists with correct structure."""
        data = load_historical_data()
        excel_path = save_historical_report_excel(data)

        wb = load_workbook(excel_path)
        assert "Summary" in wb.sheetnames

        ws = wb["Summary"]
        # Check header row
        headers = [cell.value for cell in ws[1]]
        assert "Provider" in headers
        assert "Method" in headers
        assert "Size" in headers
        assert "Avg Upload (MiBps)" in headers
        assert "Avg Download (MiBps)" in headers
        assert "Avg Latency (sec)" in headers
        assert "Dates Tested" in headers
        assert "Total Runs" in headers

    def test_has_upload_sheet(self, csv_file, temp_results_dir):
        """Test that Upload Mean sheet exists."""
        data = load_historical_data()
        excel_path = save_historical_report_excel(data)

        wb = load_workbook(excel_path)
        assert "Upload Mean (MiBps)" in wb.sheetnames

        ws = wb["Upload Mean (MiBps)"]
        headers = [cell.value for cell in ws[1]]
        assert "Date" in headers
        assert "Provider" in headers
        assert "Method" in headers

    def test_has_download_sheet(self, csv_file, temp_results_dir):
        """Test that Download Mean sheet exists."""
        data = load_historical_data()
        excel_path = save_historical_report_excel(data)

        wb = load_workbook(excel_path)
        assert "Download Mean (MiBps)" in wb.sheetnames

    def test_has_latency_sheet(self, csv_file, temp_results_dir):
        """Test that Latency Mean sheet exists."""
        data = load_historical_data()
        excel_path = save_historical_report_excel(data)

        wb = load_workbook(excel_path)
        assert "Latency Mean (sec)" in wb.sheetnames

    def test_has_raw_data_sheet(self, csv_file, temp_results_dir):
        """Test that Raw Data sheet exists."""
        data = load_historical_data()
        excel_path = save_historical_report_excel(data)

        wb = load_workbook(excel_path)
        assert "Raw Data" in wb.sheetnames

    def test_metric_sheet_has_size_columns(self, csv_file, temp_results_dir):
        """Test that metric sheets have size columns."""
        data = load_historical_data()
        excel_path = save_historical_report_excel(data)

        wb = load_workbook(excel_path)
        ws = wb["Upload Mean (MiBps)"]
        headers = [cell.value for cell in ws[1]]

        # Should have size columns
        assert "1KB" in headers
        assert "10KB" in headers

    def test_data_rows_contain_correct_values(self, csv_file, temp_results_dir):
        """Test that data rows contain correct aggregated values."""
        data = load_historical_data()
        excel_path = save_historical_report_excel(data)

        wb = load_workbook(excel_path)
        ws = wb["Upload Mean (MiBps)"]

        # Find row for 2026-03-01, aws, sdk
        headers = [cell.value for cell in ws[1]]
        date_idx = headers.index("Date") + 1
        provider_idx = headers.index("Provider") + 1
        method_idx = headers.index("Method") + 1
        kb1_idx = headers.index("1KB") + 1

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == "2026-03-01" and row[1] == "aws" and row[2] == "sdk":
                # Mean of 10.5 and 11.5 = 11.0
                assert row[kb1_idx - 1] == 11.0
                break

    def test_summary_aggregates_across_dates(self, csv_file, temp_results_dir):
        """Test that Summary sheet aggregates across all dates per size."""
        data = load_historical_data()
        excel_path = save_historical_report_excel(data)

        wb = load_workbook(excel_path)
        ws = wb["Summary"]

        # Find aws sdk 1KB row (size in column C, dates in column G)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == "aws" and row[1] == "sdk" and row[2] == "1KB":
                # Dates tested should be 2 (2026-03-01 and 2026-03-05)
                dates_idx = 6  # "Dates Tested" column (0-indexed)
                assert row[dates_idx] == 2
                break

    def test_returns_path_to_created_file(self, csv_file, temp_results_dir):
        """Test that function returns path to created file."""
        data = load_historical_data()
        excel_path = save_historical_report_excel(data)

        assert isinstance(excel_path, Path)
        assert excel_path.name.endswith("-history.xlsx")
